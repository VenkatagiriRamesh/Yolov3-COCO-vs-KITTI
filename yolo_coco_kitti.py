#Computer Vision Project - Venkatagiri Ramesh & Pranav Krishna Prasad

import numpy as np
import time
import cv2
import matplotlib.pyplot as plt
import glob
from openpyxl import Workbook,load_workbook


cocoxl_filename = 'prediction_data/coco.xlsx'
kittixl_filename = 'prediction_data/kitti.xlsx'
book_coco = Workbook()
book_kitti = Workbook()
sheet_coco = book_coco.active
sheet_kitti = book_kitti.active

filenames = glob.glob("test_images/*.jpg")
filenames.sort()
img_coco = []
img_kitti = []

'''YOLOv3_COCO - Initialization'''
labels_coco_path = "yolo_coco/coco.names"
labels_coco = open(labels_coco_path).read().strip().split("\n")

np.random.seed(42)
colors_coco = np.random.randint(0, 255, size=(len(labels_coco), 3),
	dtype="uint8")

weights_coco = "yolo_coco/yolov3.weights"
config_coco = "yolo_coco/yolov3.cfg"

net_coco = cv2.dnn.readNetFromDarknet(config_coco, weights_coco)


'''YOLOv3_KIITI- Initialization'''
labels_kitti_path= "yolo_kitti/kitti.names"
labels_kitti = open(labels_kitti_path).read().strip().split("\n")

np.random.seed(42)
colors_kitti = np.random.randint(0, 255, size=(len(labels_kitti), 3),
	dtype="uint8")

weights_kitti = "yolo_kitti/yolov3-kitti.weights"
config_kitti = "yolo_kitti/yolov3-kitti.cfg"

net_kitti = cv2.dnn.readNetFromDarknet(config_kitti, weights_kitti)
sheet_kitti.cell(row=1,column=1).value = 'Filename'
sheet_kitti.cell(row=1,column=2).value = 'Class ID'
sheet_kitti.cell(row=1,column=3).value = 'Prediction Confidence'
sheet_kitti.cell(row=1,column=4).value = 'x1'
sheet_kitti.cell(row=1,column=5).value = 'y1'
sheet_kitti.cell(row=1,column=6).value = 'x2'
sheet_kitti.cell(row=1,column=7).value = 'y2'
xl_count = 2
xl_counts = 2
ih=0
iw=0
nc = 0
nk = 0
'''Image input pipline'''
for img in filenames:
	img_coco =cv2.imread(img)
	img_kitti =cv2.imread(img)
	print(img)

	'''Image yolov3 coco processing'''
	sheet_coco.cell(row=1,column=1).value = 'Filename'
	sheet_coco.cell(row=1,column=2).value = 'Class ID'
	sheet_coco.cell(row=1,column=3).value = 'Prediction Confidence'
	sheet_coco.cell(row=1,column=4).value = 'x1'
	sheet_coco.cell(row=1,column=5).value = 'y1'
	sheet_coco.cell(row=1,column=6).value = 'x2'
	sheet_coco.cell(row=1,column=7).value = 'y2'

	(H, W) = img_coco.shape[:2]
	ih = H
	iw = W
	ln = net_coco.getLayerNames()
	ln = [ln[i[0] - 1] for i in net_coco.getUnconnectedOutLayers()]

	blob_coco = cv2.dnn.blobFromImage(img_coco, 1 / 255.0, (416, 416),swapRB=True, crop=False)
	net_coco.setInput(blob_coco)
	start_coco = time.time()
	layerOutputs = net_coco.forward(ln)
	end_coco = time.time()
	predtime_coco = end_coco - start_coco

	boxes_coco = []
	confidences_coco = []
	classIDs_coco = []
	xbar_coco = []
	ybar_coco = []
	x1_coco = []
	y1_coco = []
	x2_coco = []
	y2_coco = []
	for output in layerOutputs:
		for detection in output:
			scores = detection[5:]
			classID = np.argmax(scores)
			confidence = scores[classID]

			if confidence > 0.2:
				box = detection[0:4] * np.array([W, H, W, H])
				(centerX, centerY, width, height) = box.astype("int")

				x = int(centerX - (width / 2))
				y = int(centerY - (height / 2))

				boxes_coco.append([x, y, int(width), int(height)])
				confidences_coco.append(float(confidence))
				classIDs_coco.append(classID)
				xbar_coco.append(labels_coco[classID])
				ybar_coco.append(confidence)

	idxs = cv2.dnn.NMSBoxes(boxes_coco, confidences_coco, 0.5,0.3)
	nc = len(idxs)

	if len(idxs) > 0:
		for i in idxs.flatten():
			(x, y) = (boxes_coco[i][0], boxes_coco[i][1])
			(w, h) = (boxes_coco[i][2], boxes_coco[i][3])

			color = [int(c) for c in colors_coco[classIDs_coco[i]]]
			cv2.rectangle(img_coco, (x, y), (x + w, y + h), color, 2)
			x1_coco.append(x)
			y1_coco.append(y)
			x2_coco.append(int(x+w))
			y2_coco.append(int(y+h))

			text = "{}: {:.4f}".format(labels_coco[classIDs_coco[i]], confidences_coco[i])
			cv2.putText(img_coco, text, (x, y - 5), cv2.FONT_HERSHEY_SIMPLEX,0.5, color, 2)
			sheet_coco.cell(row=xl_count,column=1).value = img
			sheet_coco.cell(row=xl_count,column=2).value = labels_coco[classIDs_coco[i]]
			sheet_coco.cell(row=xl_count,column=3).value = confidences_coco[i]
			sheet_coco.cell(row=xl_count,column=4).value = x
			sheet_coco.cell(row=xl_count,column=5).value = y
			sheet_coco.cell(row=xl_count,column=6).value = x+w
			sheet_coco.cell(row=xl_count,column=7).value = y+h
			xl_count = xl_count + 1
			book_coco.save(cocoxl_filename)

	'''Image yolov3 kitti processing'''
	(H, W) = img_kitti.shape[:2]
	ln = net_kitti.getLayerNames()
	ln = [ln[i[0] - 1] for i in net_kitti.getUnconnectedOutLayers()]

	blob_kitti = cv2.dnn.blobFromImage(img_kitti, 1 / 255.0, (416, 416),swapRB=True, crop=False)
	net_kitti.setInput(blob_kitti)
	start_kitti = time.time()
	layerOutputs = net_kitti.forward(ln)
	end_kitti = time.time()
	predtime_kitti = end_kitti - start_kitti

	boxes_kitti = []
	confidences_kitti = []
	classIDs_kitti = []
	xbar_kitti = []
	ybar_kitti = []
	x1_kitti = []
	y1_kitti = []
	x2_kitti = []
	y2_kitti = []

	for output in layerOutputs:
		for detection in output:
			scores = detection[5:]
			classID = np.argmax(scores)
			confidence = scores[classID]

			if confidence > 0.2:
				box = detection[0:4] * np.array([W, H, W, H])
				(centerX, centerY, width, height) = box.astype("int")

				x = int(centerX - (width / 2))
				y = int(centerY - (height / 2))

				boxes_kitti.append([x, y, int(width), int(height)])
				confidences_kitti.append(float(confidence))
				classIDs_kitti.append(classID)
				xbar_kitti.append(labels_kitti[classID])
				ybar_kitti.append(confidence)

	idxs_kitti = cv2.dnn.NMSBoxes(boxes_kitti, confidences_kitti, 0.5,0.3)
	nk = len(idxs_kitti)
	if len(idxs_kitti) > 0:
		for i in idxs_kitti.flatten():
			(x, y) = (boxes_kitti[i][0], boxes_kitti[i][1])
			(w, h) = (boxes_kitti[i][2], boxes_kitti[i][3])
			x1_kitti.append(x)
			y1_kitti.append(y)
			x2_kitti.append(int(x+w))
			y2_kitti.append(int(y+h))
			color = [int(c) for c in colors_kitti[classIDs_kitti[i]]]
			cv2.rectangle(img_kitti, (x, y), (x + w, y + h), color, 2)

			text = "{}: {:.4f}".format(labels_kitti[classIDs_kitti[i]], confidences_kitti[i])
			cv2.putText(img_kitti, text, (x, y - 5), cv2.FONT_HERSHEY_SIMPLEX,0.5, color, 2)
			sheet_kitti.cell(row=xl_counts,column=1).value = img
			sheet_kitti.cell(row=xl_counts,column=2).value = labels_kitti[classIDs_kitti[i]]
			sheet_kitti.cell(row=xl_counts,column=3).value = confidences_coco[i]
			sheet_kitti.cell(row=xl_counts,column=4).value = x
			sheet_kitti.cell(row=xl_counts,column=5).value = y
			sheet_kitti.cell(row=xl_counts,column=6).value = x+w
			sheet_kitti.cell(row=xl_counts,column=7).value = y+h
			xl_counts = xl_counts + 1
			book_kitti.save(kittixl_filename)

	predtime_label = ['predtime_coco','predtime_kitti']

	pred_time = [predtime_coco,predtime_kitti]

	co1_coco = list(zip(x1_coco,y1_coco))
	co1_kitti = list(zip(x1_kitti,y1_kitti))
	co1_int = [i for i in co1_coco if i in co1_kitti]
	co1_per = len(co1_int)/(len(co1_coco)+len(co1_kitti))
	co1_per = co1_per* 100

	co2_coco = list(zip(x2_coco,y2_coco))
	co2_kitti = list(zip(x2_kitti,y2_kitti))
	co2_int = [i for i in co2_coco if i in co2_kitti]
	co2_per = len(co2_int)/(len(co2_coco)+len(co2_kitti))
	co2_per = co2_per* 100

	fig = plt.figure('YOLOv3 - COCO vs KIITI')
	ax1 = plt.subplot2grid((17, 3), (0, 0), rowspan=5)
	ax2 = plt.subplot2grid((17, 3), (6, 0), rowspan=5)
	ax3 = plt.subplot2grid((17, 3), (12, 0), rowspan=5)
	ax4 = plt.subplot2grid((17, 3), (0, 1), rowspan=8)
	ax5 = plt.subplot2grid((17, 3), (9, 1), rowspan=8)
	ax6 = plt.subplot2grid((17,3),(0,2),rowspan=5)
	ax7 = plt.subplot2grid((17,3),(6,2),rowspan=5)
	ax8 = plt.subplot2grid((17,3),(12,2),rowspan=5)


	ax1.title.set_text('YOLOv3_COCO - Class labels vs Prediction Confidence')
	ax2.title.set_text('YOLOv3_KIITI - Class labels vs Prediction Confidence')
	ax3.title.set_text('Prediction Time')
	ax4.title.set_text('YOLOv3_COCO Prediction')
	ax5.title.set_text('YOLOv3_KITTI Prediction')
	ax6.title.set_text('Bounding Box x1-y1')
	ax7.title.set_text('Bouding Box x2-y2')
	ax8.title.set_text('Bouding Box position deviation')

	ax1.bar(xbar_coco,ybar_coco, color='SkyBlue',width=0.4)
	for a,b in zip(xbar_coco, ybar_coco):
	    ax1.text(a, b, str(b))
	ax2.bar(xbar_kitti,ybar_kitti,color='SkyBlue',width=0.4)
	for a,b in zip(xbar_kitti,ybar_kitti):
	    ax2.text(a, b, str(b))
	ax3.bar(predtime_label,pred_time,color='IndianRed',width=0.7)
	for a,b in zip(predtime_label,pred_time):
	    ax3.text(a, b, str(b))
	ymin=0
	ymax=ih
	xmin = 0
	xmax = iw
	ax6.scatter(x1_coco,y1_coco,c='b',label = 'coco')
	ax6.scatter(x1_kitti,y1_kitti,c='r',label='kitti')
	ax6.set_ylim([ymax,ymin])
	ax6.set_xlim([xmin,xmax])
	ax6.legend(loc='upper right');
	ax7.scatter(x2_coco,y2_coco,c='g',label = 'coco')
	ax7.scatter(x2_kitti,y2_kitti,c='m',label='kitti')
	ax7.set_ylim([ymax,ymin])
	ax7.set_xlim([xmin,xmax])
	ax7.legend(loc='upper right');
	ax4.imshow(img_coco)
	ax5.imshow(img_kitti)
	ax8.text(0.1, 0.60, 'Deviation of x1-y1 of boudning boxes - ' + str(100.0-co1_per)+'%', style='italic',bbox={'facecolor': 'red', 'alpha': 0.5, 'pad': 3})
	ax8.text(0.1, 0.30, 'Deviation of x2-y2 of boudning boxes - ' + str(100.0-co2_per)+'%', style='italic',bbox={'facecolor': 'red', 'alpha': 0.5, 'pad': 3})

	plt.show()

	c = cv2.waitKey(0)
	if 'q' == chr(c & 255):
		continue
